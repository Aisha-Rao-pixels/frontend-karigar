"""
daily_summary_service.py
------------------------
Sends a professional Excel report to the manager every day at 5:30 PM IST.

The Excel file contains ALL workers registered from Monday of the current week
up to and including today — so Friday's email has Mon + Tue + Wed + Thu + Fri data.
The sheet resets every Monday (new week = fresh sheet starting from Monday only).

The manager gets column-filter buttons on every column so they can sort/filter
by city, skill, status, date, etc., right inside Excel — no extra tools needed.

Environment variables required (same ones already used by email_service.py):
  RESEND_API_KEY          — your Resend API key
  RESEND_FROM_EMAIL       — sender address  e.g. "Karigar <no-reply@yourdomain.com>"
  PROFILE_EMAIL_RECIPIENT — manager's email address
"""

import asyncio
import base64
import io
import logging
import os
from datetime import datetime, timezone, timedelta

import httpx
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Resend config (reuses what's already in your .env) ──────────────────────
RESEND_API_KEY    = os.environ.get("RESEND_API_KEY")
RESEND_FROM_EMAIL = os.environ.get("RESEND_FROM_EMAIL")
MANAGER_EMAIL     = os.environ.get("PROFILE_EMAIL_RECIPIENT")
RESEND_API_URL    = "https://api.resend.com/emails"

# ── IST = UTC + 5:30 ────────────────────────────────────────────────────────
IST_OFFSET = timezone(timedelta(hours=5, minutes=30))

# ── Brand colours ────────────────────────────────────────────────────────────
BRAND_DARK   = "7A2E1D"   # dark red  — header background
BRAND_MID    = "C0392B"   # medium red
BRAND_LIGHT  = "FAF0EE"   # very light red — alternate row
WHITE        = "FFFFFF"
BLACK        = "000000"
MUTED        = "6B6B6B"
GREEN_BG     = "E6F4EA"   # approved status fill
ORANGE_BG    = "FFF3E0"   # pending  status fill
RED_BG       = "FFEBEE"   # rejected status fill
HEADER_FONT  = "FFFFFF"

# ── Column definitions ───────────────────────────────────────────────────────
# (header_label, worker_field_or_callable, column_width)
COLUMNS = [
    ("S.No",             None,                          6),
    ("Full Name",        "full_name",                   22),
    ("Phone",            "phone",                       14),
    ("Gender",           "gender",                      10),
    ("City",             "city",                        14),
    ("Area",             "area",                        16),
    ("Skills",           "skills",                      28),
    ("Experience (yrs)", "years_experience",            16),
    ("Languages",        "languages",                   18),
    ("Current Employer", "current_employer",            22),
    ("Prev. Employer",   "previous_employer",           22),
    ("Wage (₹/month)",   "wage_expectation",            14),
    ("Availability",     "availability_status",         16),
    ("Verification",     "verification_status",         14),
    ("Referral Code",    "referral_code",               14),
    ("Referred By",      "referred_by_code",            18),
    ("Registered On",    "created_at",                  18),
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def _avail_label(val: str) -> str:
    return {
        "available_now":  "Available Now",
        "available_from": "Available From Date",
        "not_available":  "Not Available",
    }.get(val or "", val or "—")


def _verify_label(val: str) -> str:
    return {
        "approved": "✓ Approved",
        "pending":  "⏳ Pending",
        "rejected": "✗ Rejected",
    }.get(val or "", val or "—")


def _verify_fill(val: str) -> PatternFill | None:
    colour = {
        "approved": GREEN_BG,
        "pending":  ORANGE_BG,
        "rejected": RED_BG,
    }.get(val or "")
    return PatternFill("solid", fgColor=colour) if colour else None


def _cell_value(worker: dict, field) -> str:
    if field is None:
        return ""
    raw = worker.get(field)
    if raw is None or raw == "":
        return "—"
    if isinstance(raw, list):
        return ", ".join(raw) if raw else "—"
    if field == "phone":
        return f"+91 {raw}"
    if field == "availability_status":
        return _avail_label(raw)
    if field == "verification_status":
        return _verify_label(raw)
    if field == "created_at":
        # ISO string → readable date
        try:
            return raw[:10]
        except Exception:
            return str(raw)
    return str(raw)


def _thin_border() -> Border:
    s = Side(style="thin", color="D0C4BE")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Excel builder ────────────────────────────────────────────────────────────

def build_excel(workers: list[dict], week_start: datetime, today: datetime) -> bytes:
    """
    Build and return the Excel bytes for a weekly rolling report.
    `workers` — list of worker dicts for Mon–today.
    `week_start` — the Monday datetime (IST).
    `today`      — today's datetime (IST).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Registrations"

    # ── Title row (row 1) ────────────────────────────────────────────────────
    date_range = (
        f"{week_start.strftime('%d %b %Y')} – {today.strftime('%d %b %Y')}"
    )
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"Karigar — User Registration Report  |  {date_range}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color=HEADER_FONT)
    title_cell.fill = PatternFill("solid", fgColor=BRAND_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sub-title row (row 2) ────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"Total registrations this week: {len(workers)}   |   "
        f"Generated: {today.strftime('%d %b %Y, %I:%M %p')} IST"
    )
    sub_cell.font = Font(name="Arial", size=9, color=MUTED)
    sub_cell.fill = PatternFill("solid", fgColor="F5EDE9")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Column header row (row 3) ────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=BRAND_MID)
    header_font = Font(name="Arial", bold=True, size=10, color=HEADER_FONT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 30

    # ── Data rows (starting row 4) ──────────────────────────────────────────
    for row_num, worker in enumerate(workers, start=1):
        excel_row = row_num + 3           # offset by 3 header rows
        is_alt    = row_num % 2 == 0      # alternating row colour
        row_fill  = PatternFill("solid", fgColor=BRAND_LIGHT) if is_alt else None
        data_font = Font(name="Arial", size=9, color=BLACK)
        data_align_left   = Alignment(vertical="center", wrap_text=True)
        data_align_center = Alignment(horizontal="center", vertical="center")

        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            if col_idx == 1:
                # S.No
                value = row_num
            else:
                value = _cell_value(worker, field)

            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = _thin_border()

            # Special fill for verification status column (col 14)
            if field == "verification_status":
                v_fill = _verify_fill(worker.get("verification_status", ""))
                cell.fill = v_fill or (row_fill or PatternFill())
                cell.alignment = data_align_center
            else:
                if row_fill:
                    cell.fill = row_fill
                cell.alignment = data_align_left if col_idx > 2 else data_align_center

        ws.row_dimensions[excel_row].height = 18

    # ── Auto-filter (Excel filter dropdowns on every column header) ─────────
    ws.auto_filter.ref = (
        f"A3:{get_column_letter(len(COLUMNS))}{3 + len(workers)}"
    )

    # ── Freeze top 3 rows so headers stay visible while scrolling ───────────
    ws.freeze_panes = "A4"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    summary_title = ws2.cell(row=1, column=1, value="Weekly Summary")
    summary_title.font = Font(name="Arial", bold=True, size=12, color=HEADER_FONT)
    summary_title.fill = PatternFill("solid", fgColor=BRAND_DARK)
    ws2.merge_cells("A1:B1")
    summary_title.alignment = Alignment(horizontal="center")

    def s_row(row, label, value):
        lc = ws2.cell(row=row, column=1, value=label)
        vc = ws2.cell(row=row, column=2, value=value)
        lc.font = Font(name="Arial", bold=True, size=10)
        vc.font = Font(name="Arial", size=10)
        lc.border = vc.border = _thin_border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    approved = sum(1 for w in workers if w.get("verification_status") == "approved")
    pending  = sum(1 for w in workers if w.get("verification_status") == "pending")
    rejected = sum(1 for w in workers if w.get("verification_status") == "rejected")

    s_row(2, "Report Period",          date_range)
    s_row(3, "Total New Users",        len(workers))
    s_row(4, "Approved",               approved)
    s_row(5, "Pending Verification",   pending)
    s_row(6, "Rejected",               rejected)
    s_row(7, "Generated At (IST)",     today.strftime("%d %b %Y, %I:%M %p"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Email sender ─────────────────────────────────────────────────────────────

async def send_daily_summary(db) -> bool:
    """
    Build the weekly-rolling Excel report and email it to the manager.
    Called by the scheduler every day at 5:30 PM IST, and also available
    as a manual trigger via the admin endpoint.
    """
    if not (RESEND_API_KEY and RESEND_FROM_EMAIL and MANAGER_EMAIL):
        logger.warning("Daily summary: email not configured (missing RESEND env vars).")
        return False

    try:
        now_ist   = datetime.now(IST_OFFSET)
        today_ist = now_ist.date()

        # Monday of the current week (weekday 0 = Monday)
        days_since_monday = today_ist.weekday()           # 0 = Mon, 6 = Sun
        monday            = today_ist - timedelta(days=days_since_monday)

        # Build ISO strings for DB query  (Monday 00:00:00 → today 23:59:59 IST)
        monday_start_utc = (
            datetime(monday.year, monday.month, monday.day, 0, 0, 0, tzinfo=IST_OFFSET)
            .astimezone(timezone.utc)
            .isoformat()
            .replace("+00:00", "Z")
        )
        today_end_utc = (
            datetime(today_ist.year, today_ist.month, today_ist.day, 23, 59, 59, tzinfo=IST_OFFSET)
            .astimezone(timezone.utc)
            .isoformat()
            .replace("+00:00", "Z")
        )

        # Fetch workers registered Mon–today
        workers = await db.workers.find({
            "created_at": {
                "$gte": monday_start_utc,
                "$lte": today_end_utc,
            }
        }).sort("created_at", 1).to_list(10000)

        logger.info(
            "Daily summary: found %d workers from %s to %s",
            len(workers), monday, today_ist,
        )

        week_start_dt = datetime(monday.year, monday.month, monday.day, tzinfo=IST_OFFSET)
        excel_bytes   = build_excel(workers, week_start_dt, now_ist)

        # ── Compose email ────────────────────────────────────────────────────
        day_name     = now_ist.strftime("%A")           # e.g. "Friday"
        date_str     = now_ist.strftime("%d %B %Y")     # e.g. "04 July 2025"
        week_range   = f"{monday.strftime('%d %b')} – {today_ist.strftime('%d %b %Y')}"
        filename     = f"Karigar_Users_{monday.strftime('%d%b')}_{today_ist.strftime('%d%b%Y')}.xlsx"

        subject = (
            f"Karigar Daily Update — {len(workers)} users this week ({date_str})"
        )

        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:auto;">
          <div style="background:#7A2E1D;padding:24px 28px;border-radius:8px 8px 0 0;">
            <h2 style="color:#fff;margin:0;font-size:20px;">Karigar — Daily Report</h2>
            <p style="color:#F5EDE9;margin:6px 0 0;font-size:13px;">{day_name}, {date_str}</p>
          </div>
          <div style="background:#fff;padding:24px 28px;border:1px solid #e8e0dc;border-top:none;border-radius:0 0 8px 8px;">
            <p style="font-size:14px;color:#333;">Good evening,</p>
            <p style="font-size:14px;color:#333;">
              I'm your Karigar AI agent. Here's your daily update for the week 
              <strong>{week_range}</strong>.
            </p>
            <div style="background:#FAF0EE;border-left:4px solid #7A2E1D;padding:14px 18px;border-radius:4px;margin:18px 0;">
              <p style="margin:0;font-size:22px;font-weight:bold;color:#7A2E1D;">{len(workers)}</p>
              <p style="margin:4px 0 0;font-size:13px;color:#6B6B6B;">users registered so far this week</p>
            </div>
            <p style="font-size:13px;color:#333;">
              The full user list (with all details) is attached as an Excel file.<br>
              You can open it in Excel or Google Sheets and use the <strong>filter dropdowns</strong>
              at the top of each column to sort or filter by city, skill, status, and more.
            </p>
            <p style="font-size:12px;color:#999;margin-top:24px;border-top:1px solid #eee;padding-top:12px;">
              This report is generated automatically every day at 5:30 PM IST.<br>
              The Excel sheet contains data from Monday ({monday.strftime('%d %b')}) through today ({today_ist.strftime('%d %b')}).
              It resets every Monday with a fresh week.
            </p>
          </div>
        </div>
        """

        payload = {
            "from": RESEND_FROM_EMAIL,
            "to":   [MANAGER_EMAIL],
            "subject": subject,
            "html":    html,
            "attachments": [{
                "filename": filename,
                "content":  base64.b64encode(excel_bytes).decode("ascii"),
            }],
        }

        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                RESEND_API_URL,
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type":  "application/json",
                },
                json=payload,
            )

        if resp.status_code >= 400:
            logger.error("Resend error: %s %s", resp.status_code, resp.text)
            return False

        logger.info(
            "Daily summary sent to %s — %d workers, week %s",
            MANAGER_EMAIL, len(workers), week_range,
        )
        return True

    except Exception as exc:
        logger.error("Daily summary failed: %s", exc, exc_info=True)
        return False
